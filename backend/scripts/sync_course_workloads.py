from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import inspect, text

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app import create_app, db  # noqa: E402
from app.models import Course, Department, Program  # noqa: E402


def parse_args():
    parser = argparse.ArgumentParser(description="Sync course workloads from an Excel workbook.")
    parser.add_argument(
        "--workbook",
        default=str(ROOT_DIR.parent / "Work Load CSE & IT SPRING 2025.xlsx"),
        help="Path to the Excel workbook",
    )
    parser.add_argument(
        "--sheet",
        default="Course -wise",
        help="Workbook sheet name",
    )
    return parser.parse_args()


def ensure_course_workload_columns():
    inspector = inspect(db.engine)
    existing_columns = {column["name"] for column in inspector.get_columns("course")}
    missing = [
        column_name
        for column_name in ("lecture_hours", "tutorial_hours", "practical_hours")
        if column_name not in existing_columns
    ]

    for column_name in missing:
        db.session.execute(
            text(f"ALTER TABLE course ADD COLUMN {column_name} INTEGER NOT NULL DEFAULT 0")
        )

    if missing:
        db.session.commit()


def normalize_string(value):
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def parse_int(value):
    if value is None or str(value).strip() == "":
        return 0
    return max(0, int(float(value)))


def parse_program_semester(raw_program):
    normalized = normalize_string(raw_program)
    if not normalized:
        return None, None

    parts = normalized.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return parts[0].strip(), int(parts[1])
    return normalized, None


def infer_course_type(lecture_hours, tutorial_hours, practical_hours):
    if practical_hours > 0 and lecture_hours == 0 and tutorial_hours == 0:
        return "Lab"
    return "Theory"


def get_header_row_index(sheet):
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        first = normalize_string(row[0] if row else None)
        second = normalize_string(row[1] if row and len(row) > 1 else None)
        if first == "Program" and second == "Course Code":
            return index
    raise RuntimeError("Could not find the workbook header row")


def resolve_department(program_code):
    program = Program.query.filter_by(code=program_code).first()
    if not program or not program.department_id:
        return None
    return db.session.get(Department, program.department_id)


def iter_workbook_rows(workbook_path, sheet_name):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    header_row_index = get_header_row_index(sheet)

    header = next(
        sheet.iter_rows(
            min_row=header_row_index,
            max_row=header_row_index,
            values_only=True,
        )
    )
    headers = [normalize_string(value) for value in header]
    current_program = None

    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not any(value is not None and str(value).strip() for value in row):
            continue

        record = dict(zip(headers, row))
        program_cell = normalize_string(record.get("Program"))
        if program_cell:
            current_program = program_cell
        else:
            record["Program"] = current_program

        yield record


def sync_workloads(workbook_path, sheet_name):
    ensure_course_workload_columns()

    updated = 0
    created = 0
    skipped = []

    for row_number, record in enumerate(iter_workbook_rows(workbook_path, sheet_name), start=1):
        code = normalize_string(record.get("Course Code"))
        name = normalize_string(record.get("Course"))
        program_code, semester = parse_program_semester(record.get("Program"))
        lecture_hours = parse_int(record.get("L"))
        tutorial_hours = parse_int(record.get("T"))
        practical_hours = parse_int(record.get("P"))

        if not code:
            skipped.append(f"Row {row_number}: skipped because Course Code is blank ({name or 'Unnamed course'})")
            continue

        if not name:
            skipped.append(f"Row {row_number}: skipped because Course name is blank for code {code}")
            continue

        if not program_code:
            skipped.append(f"Row {row_number}: skipped because Program is blank for code {code}")
            continue

        department = resolve_department(program_code)
        if not department:
            skipped.append(f"Row {row_number}: skipped because program '{program_code}' is not mapped to a department")
            continue

        course = Course.query.filter_by(code=code).first()
        is_new = course is None
        if is_new:
            course = Course(code=code)
            db.session.add(course)

        course.name = name
        course.code = code
        course.program_code = program_code
        course.semester = semester
        course.semester_name = f"Semester {semester}" if semester is not None else None
        course.course_type = infer_course_type(lecture_hours, tutorial_hours, practical_hours)
        course.department_id = department.id
        course.department_code = department.code
        course.lecture_hours = lecture_hours
        course.tutorial_hours = tutorial_hours
        course.practical_hours = practical_hours

        if is_new:
            created += 1
        else:
            updated += 1

    db.session.commit()
    return updated, created, skipped


def main():
    args = parse_args()
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    app = create_app("development")
    with app.app_context():
        updated, created, skipped = sync_workloads(workbook_path, args.sheet)
        print(f"Updated courses: {updated}")
        print(f"Created courses: {created}")
        print(f"Skipped rows: {len(skipped)}")
        for item in skipped:
            print(item)


if __name__ == "__main__":
    main()
